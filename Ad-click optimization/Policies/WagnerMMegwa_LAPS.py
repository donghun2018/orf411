
import numpy as np
from .policy import Policy
from openpyxl import load_workbook

class Policy_WagnerMMegwa_LAPS(Policy):

    def __init__(self, all_attrs, possible_bids=list(range(10)), max_t=10, randseed=12345):
        """
        initializes policy base class.
        :param all_attrs: list of all possible attributes
        :param possible_bids: list of all allowed bids
        :param max_t: maximum number of auction 'iter', or iteration timesteps
        :param randseed: random number seed.
        """
        super().__init__(all_attrs, possible_bids, max_t, randseed=randseed)

        self.initialize_belief_model()

    def initialize_belief_model(self):
        
        nbid = len(self.bid_space)
        nattr = len(self.attrs)
        self.alp = np.array([[9,6,7,27,0,2,20,0,0,0], [45,0,0,39,0,0,14,6,1,0], [21,0,22,1,0,18,1,0,0,0], [9,6,4,5,5,7,25,0,0,0], [5,4,3,8,5,5,0,8,0,0], [9,20,1,25,16,0,1,0,0,0], [9,3,13,7,8,18,0,0,4,0], [5,2,5,14,1,8,4,0,5,0], [17,0,4,9,1,0,22,0,0,0],[3,5,12,1,0,0,9,0,2,0],[15,0,13,4,28,1,0,1,0,0],[5,4,4,15,3,0,0,0,5,0],[2,3,1,1,11,0,0,0,0,0],[13,1,11,20,1,0,0,0,0,0]])
        self.bet = np.array([[11,10,23,19,0,1,27,5,1,0],[22,0,0,24,0,2,12,3,0,0],[19,1,34,0,3,44,3,0,0,0],[10,9,15,8,12,14,32,3,1,3],[13,13,10,20,5,26,0,41,0,0],[18,22,0,32,21,3,0,0,0,0],[11,5,13,14,11,30,0,0,20,0],[10,8,9,37,1,17,8,0,25,0],[41,2,8,18,2,0,41,0,0,0],[16,12,69,1,0,0,17,0,16,0],[40,0,12,6,46,1,0,1,0,0],[10,17,15,64,4,0,0,0,16,0],[14,19,23,1,83,0,0,4,0,0],[34,1,23,55,6,3,0,0,0,0]])
        self.alphas = np.zeros((nattr, nbid))
        self.betas = np.zeros((nattr, nbid))
        
        for i in range (nattr):
            for j in range (nbid):
                if j < 10:
                    self.alphas[i][j] = self.alp[i][0]
                    self.betas[i][j] = self.bet[i][0]
                elif j < 20:
                    self.alphas[i][j] = self.alp[i][1]
                    self.betas[i][j] = self.bet[i][1]
                elif j < 30:
                    self.alphas[i][j] = self.alp[i][2]
                    self.betas[i][j] = self.bet[i][2]
                elif j < 40:
                    self.alphas[i][j] = self.alp[i][3]
                    self.betas[i][j] = self.bet[i][3]
                elif j < 50:
                    self.alphas[i][j] = self.alp[i][4]
                    self.betas[i][j] = self.bet[i][4]
                elif j < 60:
                    self.alphas[i][j] = self.alp[i][5]
                    self.betas[i][j] = self.bet[i][5]
                elif j < 70:
                    self.alphas[i][j] = self.alp[i][6]
                    self.betas[i][j] = self.bet[i][6]
                elif j < 80:
                    self.alphas[i][j] = self.alp[i][7]
                    self.betas[i][j] = self.bet[i][7]
                elif j < 90:
                    self.alphas[i][j] = self.alp[i][8]
                    self.betas[i][j] = self.bet[i][8]
                elif j < 100:
                    self.alphas[i][j] = self.alp[i][9]
                    self.betas[i][j] = self.bet[i][9]
                




        self.var = np.zeros((nattr, nbid))
        self.probs = np.zeros((nattr, nbid))
        alpha_max = 120
        beta_max = 180
        

        #workbook = load_workbook('output_policy_info_random.xlsx')
        #sheet = workbook.active
        # counter = 0
        # while counter < 168:
        #     for i in range(2,15):
        #         for col_cells in sheet.iter_cols(min_col=7, max_col=10):
        #             num_click = int(sheet.cell(row = i + (counter*14), column = 7).value)
        #             cost_per_click = int(sheet.cell(row = i + (counter*14), column = 8).value)
        #             num_conversion = int(sheet.cell(row = i + (counter*14), column = 9).value)
        #             rev_conversion = int(sheet.cell(row = i + (counter*14), column = 10).value)
        #             update = sheet.cell(row = i, column = 4).value * 10
        #             if num_click * cost_per_click > num_conversion * rev_conversion:
        #                 self.betas[i][update] = self.betas[i][update] + 1
        #             else:
        #                 self.alphas[i][update] = self.alphas[i][update] + 1 
        #         counter = counter + 1



        for i in range(nattr):
            for j in range(nbid):
                #self.alphas[i][j] = np.floor(np.random.rand()*alpha_max)
                #self.betas[i][j] = np.floor(np.random.rand()*beta_max)
                self.var[i][j] = (self.alphas[i][j] / (self.alphas[i][j] + self.betas[i][j])) * (1 - (self.alphas[i][j] / (self.alphas[i][j] + self.betas[i][j])))

        for i in range(nattr):
            for j in range(nbid):
                self.probs[i][j] = self.alphas[i][j] / (self.alphas[i][j] + self.betas[i][j] + 1) # need to ensure the denominator is non-zero!


    def bid(self, attr):
        """
        returns a random bid, regardless of attribute
        Note how this uses internal pseudorandom number generator (self.prng) instead of np.random
        :param attr: attribute tuple. guaranteed to be found in self.attrs
        :return: a value that is found in self.bid_space
        """

        a_ix = self.attrs.index(attr)
        nbid = len(self.bid_space)

        cmax = np.zeros(nbid)

        for i in range(nbid):
            new_array = self.probs[a_ix].copy()
            new_array[i] = 0
            cmax[i] = np.amax(new_array)

        vkg = np.zeros(nbid)
        tune = 47

       

        for j in range(nbid):

            if self.alphas[a_ix][j] / (self.alphas[a_ix][j] + self.betas[a_ix][j]) < cmax[j] < (self.alphas[a_ix][j] + tune) / ((self.alphas[a_ix][j] + self.betas[a_ix][j]) + tune):
                #print("here")
                vkg[j] = self.alphas[a_ix][j] / (self.alphas[a_ix][j] + self.betas[a_ix][j]) * ((self.alphas[a_ix][j] + tune) / (self.alphas[a_ix][j] + self.betas[a_ix][j] + tune) - cmax[j])
            elif self.alphas[a_ix][j] / (self.alphas[a_ix][j] + self.betas[a_ix][j] + tune) < cmax[j] < self.alphas[a_ix][j] / (self.alphas[a_ix][j] + self.betas[a_ix][j]):
                #print("here2")
                vkg[j] = self.betas[a_ix][j] / (self.alphas[a_ix][j] + self.betas[a_ix][j]) * (cmax[j] - self.alphas[a_ix][j]) / (self.alphas[a_ix][j] + self.betas[a_ix][j] + tune)
            else:
                vkg[j] = 0

        online = 0.5

        vkgol = np.zeros(nbid)

        for i in range(nbid):
            vkgol[i] = self.probs[a_ix, i] + online * vkg[i]


        #randNum = round(np.random.uniform(0.1, 0.5), 1)
        
        choice = np.argmax(vkgol)
        #return choice

        return self.bid_space[choice] # need to return the actual bid, not the index!

    def learn(self, info):
        """
        learns from auctions results
        This policy does not learn (need not learn, because it just bids randomly)
        :param info: list of results. Single result is an aggregate outcome for all auctions with a particular attr.
                     Follows the same format as output_policy_info_?????.xlsx
        :return: does not matter
        """
        for result in info:
            a_ix = self.attrs.index(result['attr'])
            bid_index = self.bid_space.index(result['your_bid'])
            if result['cost_per_click'] == '' and result['revenue_per_conversion'] * result['num_conversion'] < result['num_click'] * result['cost_per_click']:
                self.betas[a_ix][bid_index] += 1
            else:
                self.alphas[a_ix][bid_index] += 1
        return







